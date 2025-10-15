# components/excel_generator.py
# PRODUCTION READY - Matches AllWave AV company format (Final Version)

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from io import BytesIO
import re
from datetime import datetime

# Import the image generator with better error handling
try:
    from components.product_image_generator import generate_product_info_card, extract_display_size
    IMAGE_GENERATION_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Product image generator not available: {e}")
    IMAGE_GENERATION_AVAILABLE = False
    
    # Fallback functions
    def generate_product_info_card(*args, **kwargs):
        print("⚠️ Image generation skipped (module not available)")
        return None
    
    def extract_display_size(name):
        import re
        match = re.search(r'(\d{2,3})["\']', str(name))
        return int(match.group(1)) if match else None


# ==================== STYLE DEFINITIONS ====================
def _define_styles():
    """Defines all necessary styles for the professional report."""
    thin_border_side = Side(style='thin')
    thin_border = Border(
        left=thin_border_side, 
        right=thin_border_side, 
        top=thin_border_side, 
        bottom=thin_border_side
    )
    
    return {
        "header_green_fill": PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid"),
        "header_light_green_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "table_header_blue_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "boq_category_fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "black_bold_font": Font(color="000000", bold=True),
        "bold_font": Font(bold=True),
        "thin_border": thin_border,
        "currency_format": "₹#,##0.00"
    }


# ==================== HEADER WITH LOGOS ====================
def _add_image_to_cell(sheet, image_path, cell, height_px):
    """Adds a logo to a cell, preserving aspect ratio."""
    try:
        img = ExcelImage(image_path)
        img.height = height_px
        img.width = (img.width / img.height) * height_px
        sheet.add_image(img, cell)
    except FileNotFoundError:
        # Graceful fallback - just put text
        sheet[cell] = f"Logo: {image_path}"
    except Exception as e:
        sheet[cell] = "Logo"


def _create_sheet_header(sheet):
    """Creates the standard header with four logos."""
    sheet.row_dimensions[1].height = 50
    sheet.row_dimensions[2].height = 50
    
    # Define max column for merging based on sheet title
    if "BOQ" in sheet.title:
        max_col_letter_range = 'O'
        logo_cols = {'L1:M2': 'iso_logo.png', 'N1:O2': 'avixa_logo.png'}
    else:
        max_col_letter_range = 'G'
        logo_cols = {'D1:E2': 'iso_logo.png', 'F1:G2': 'avixa_logo.png'}

    # Standard logos
    sheet.merge_cells('A1:B2')
    _add_image_to_cell(sheet, 'assets/company_logo.png', 'A1', 95)
    
    # Dynamic logos based on sheet type
    sheet.merge_cells('C1:D2')
    _add_image_to_cell(sheet, 'assets/crestron_logo.png', 'C1', 95)
    
    for cells, logo in logo_cols.items():
        if get_column_letter(sheet.max_column) >= cells.split(':')[1][0]:
            sheet.merge_cells(cells)
            _add_image_to_cell(sheet, f'assets/{logo}', cells.split(':')[0], 95)


# ==================== VERSION CONTROL SHEET ====================
def _add_version_control_sheet(workbook, project_details, styles):
    """Creates the Version Control & Contact Details sheet."""
    sheet = workbook.create_sheet(title="Version Control", index=0)
    sheet.sheet_view.showGridLines = False

    # Set column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['D'].width = 5
    sheet.column_dimensions['E'].width = 25
    sheet.column_dimensions['F'].width = 35
    
    _create_sheet_header(sheet)
    
    # === VERSION CONTROL TABLE ===
    sheet.merge_cells('A3:B3')
    vc_header = sheet['A3']
    vc_header.value = "Version"
    vc_header.fill = styles['header_green_fill']
    vc_header.font = styles['black_bold_font']
    vc_header.alignment = Alignment(horizontal='center', vertical='center')
    vc_header.border = styles['thin_border']
    sheet['B3'].border = styles['thin_border']

    vc_data = [
        ("Date of First Draft", datetime.now().strftime("%d-%b-%Y")),
        ("Date of Final Draft", ""),
        ("Version No.", "1.0"),
        ("Published Date", datetime.now().strftime("%d-%b-%Y"))
    ]
    
    for i, (label, value) in enumerate(vc_data):
        row = i + 4
        for col_letter in ['A', 'B']:
            sheet[f'{col_letter}{row}'].border = styles['thin_border']
        sheet[f'A{row}'].value = label
        sheet[f'A{row}'].fill = styles['header_light_green_fill']
        sheet[f'A{row}'].alignment = Alignment(vertical='center')
        sheet[f'B{row}'].value = value
        sheet[f'B{row}'].alignment = Alignment(vertical='center')

    # === CONTACT DETAILS TABLE ===
    sheet.merge_cells('E3:F3')
    cd_header = sheet['E3']
    cd_header.value = "Contact Details"
    cd_header.fill = styles['header_green_fill']
    cd_header.font = styles['black_bold_font']
    cd_header.alignment = Alignment(horizontal='center', vertical='center')
    cd_header.border = styles['thin_border']
    sheet['F3'].border = styles['thin_border']

    contact_data = [
        ("Design Engineer", project_details.get("Design Engineer", "")),
        ("Account Manager", project_details.get("Account Manager", "")),
        ("Client Name", project_details.get("Client Name", "")),
        ("Key Client Personnel", project_details.get("Key Client Personnel", "")),
        ("Location", project_details.get("Location", "")),
        ("PSNI Referral", "✅ YES" if project_details.get("PSNI Referral") == "Yes" else "No"),
        ("Client Type", project_details.get("Client Type", "International")),
        ("Key Comments for this version", project_details.get("Key Comments", ""))
    ]
    
    for i, (label, value) in enumerate(contact_data):
        row = i + 4
        for col_letter in ['E', 'F']:
            sheet[f'{col_letter}{row}'].border = styles['thin_border']
        sheet[f'E{row}'].value = label
        sheet[f'E{row}'].fill = styles['header_light_green_fill']
        sheet[f'E{row}'].alignment = Alignment(vertical='center')
        sheet[f'F{row}'].value = value
        if label == "Key Comments for this version":
            sheet.row_dimensions[row].height = 40
            sheet[f'F{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        else:
            sheet[f'F{row}'].alignment = Alignment(vertical='center')


# ==================== TERMS & CONDITIONS SHEET ====================
def _add_terms_and_conditions_sheet(workbook, styles):
    """
    Creates comprehensive Terms & Conditions sheet based on AllWave AV standard format.
    """
    sheet = workbook.create_sheet(title="Terms & Conditions")
    sheet.sheet_view.showGridLines = False
    _create_sheet_header(sheet)
    
    for col in get_column_letter(sheet.max_column):
        sheet.column_dimensions[col].width = 20
    
    row_cursor = 4
    
    def create_section_header(value):
        nonlocal row_cursor
        sheet.merge_cells(f'A{row_cursor}:F{row_cursor}')
        cell = sheet[f'A{row_cursor}']
        cell.value = value
        cell.fill = styles['table_header_blue_fill']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 7): sheet.cell(row=row_cursor, column=c).border = styles['thin_border']
        row_cursor += 1

    def create_terms(terms_list):
        nonlocal row_cursor
        for term in terms_list:
            sheet.merge_cells(f'A{row_cursor}:F{row_cursor}')
            cell = sheet[f'A{row_cursor}']
            cell.value = term
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if term and not term.startswith(('•', ' ')): cell.font = styles['bold_font']
            for c in range(1, 7): sheet.cell(row=row_cursor, column=c).border = styles['thin_border']
            row_cursor += 1
        row_cursor += 1
    
    # Main Title
    sheet.merge_cells(f'A{row_cursor}:F{row_cursor}')
    title_cell = sheet[f'A{row_cursor}']
    title_cell.value = "Commercial Terms & Conditions"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[row_cursor].height = 25
    for c in range(1, 7): sheet.cell(row=row_cursor, column=c).border = styles['thin_border']
    row_cursor += 2
    
    create_section_header("A. Delivery, Installations & Site Schedule")
    create_terms([
        "All Wave AV Systems undertakes to ensure its best efforts to complete the assignment within the shortest timelines possible.", "",
        "Project Schedule:", "• Week 1-3: All Wave AV Systems Design & Procurement / Client Site Preparations",
        "• Implementation: Within 12 weeks of advance payment receipt", "", "Delivery Terms:",
        "• All deliveries within 6-8 weeks of commercially clear Purchase Order", "• Equipment delivered in phased manner (max 3 shipments)", "",
        "Note: Delay in advance payment may alter project schedule. Beyond 12 weeks delay due to site issues: ₹8,000 + GST per day charge applies."
    ])
    
    create_section_header("B. Payment Terms")
    create_terms([
        "Schedule of Payment:", "• Equipment & Materials: 20% Advance with PO",
        "• Installation & Commissioning: Against system installation", "• Balance Payment: Within 30 days of ATP sign-off"
    ])

    create_section_header("C. Offer Validity")
    create_terms(["Offer Valid for 30 Days from date of quotation"])

    create_section_header("D. Placing a Purchase Order")
    create_terms([
        "Order should be placed on:", "All Wave AV Systems Pvt. Ltd.", "420A Shah & Nahar Industrial Estate,",
        "Lower Parel West, Mumbai 400013, INDIA", "", "GST No: [To be provided]", "PAN No: [To be provided]"
    ])

    create_section_header("E. Cable Estimates")
    create_terms([
        "Provisional cable estimate provided. Actual consumption may vary based on finalized layouts.",
        "Invoicing based on actual consumption: Physical measurement + 10% (for bends, curves, termination, wastage)"
    ])

    create_section_header("F. Order Changes")
    create_terms([
        "All Wave AV Systems accommodates scope changes as needed.",
        "Changes may require additional resources/time - separate Change Order will be issued.",
        "All Change Orders must be in writing with adjusted price, schedule, and acceptance criteria."
    ])
    
    create_section_header("G. Restocking / Cancellation Fees")
    create_terms(["Cancellation may involve charges up to 50% restocking/cancellation fees + shipping costs"])

    create_section_header("H. Warranty")
    create_terms([
        "All Wave AV Systems provides:", "• Comprehensive 12-month warranty on all equipment from handover date",
        "• Limited warranty on consumables (Projector lamps: 450 hours or 90 days, whichever earlier)",
        "• Extended warranty available via separate Maintenance Contract", "", "Warranty exclusions:",
        "• Power-related damage (equipment must use stabilized power/online UPS)",
        "• Accident, misuse, neglect, alteration, or component substitution", "• Fire, flood, weather exposure, force majeure events"
    ])


# ==================== ROOM BOQ SHEET (FINAL FORMAT) ====================
def _populate_room_boq_sheet(sheet, items, room_name, styles, usd_to_inr_rate, gst_rates):
    """ Creates detailed BOQ sheet with final format, colors, and column order. """
    _create_sheet_header(sheet)
    
    # Room Info Section
    info_data = [("Room Name / Room Type", room_name), ("Floor", "-"), ("Number of Seats", "-"), ("Number of Rooms", "-")]
    for i, (label, value) in enumerate(info_data):
        row = i + 3
        sheet[f'A{row}'].value, sheet[f'A{row}'].font, sheet[f'A{row}'].fill = label, styles['bold_font'], styles['header_light_green_fill']
        sheet.merge_cells(f'B{row}:C{row}')
        sheet[f'B{row}'].value = value
        for col in ['A', 'B', 'C']: sheet[f'{col}{row}'].border = styles['thin_border']
    sheet.append([])

    # Table Headers
    headers1 = ['Sr. No.', 'Description of Goods / Services', 'Make', 'Model No.', 'Qty.', 'Unit Rate (INR)', 'Total', 'Warranty', 'SGST (In Maharashtra)', None, 'CGST (In Maharashtra)', None, 'Total (TAX)', 'Total Amount (INR)', 'Reference Image']
    headers2 = [None, None, None, None, None, None, None, None, 'Rate', 'Amt', 'Rate', 'Amt', None, None, None]
    sheet.append(headers1)
    sheet.append(headers2)
    header_start_row = sheet.max_row - 1
    sheet.merge_cells(f'I{header_start_row}:J{header_start_row}')
    sheet.merge_cells(f'K{header_start_row}:L{header_start_row}')
    for row in sheet.iter_rows(min_row=header_start_row, max_row=sheet.max_row):
        for cell in row:
            if cell.value is not None:
                cell.fill, cell.font, cell.alignment, cell.border = styles["table_header_blue_fill"], Font(bold=True, color="FFFFFF"), Alignment(horizontal='center', vertical='center', wrap_text=True), styles['thin_border']
    
    # Group items and add to sheet
    grouped_items = {}
    for item in items: grouped_items.setdefault(item.get('category', 'General AV'), []).append(item)
    total_before_gst_hardware, item_s_no = 0, 1
    category_letters = [chr(ord('A') + i) for i in range(len(grouped_items))]

    for i, (category, cat_items) in enumerate(grouped_items.items()):
        sheet.append([category_letters[i], category])
        cat_row_idx = sheet.max_row
        sheet.merge_cells(f'B{cat_row_idx}:O{cat_row_idx}')
        for cell in sheet[cat_row_idx]: cell.fill, cell.font, cell.border = styles['boq_category_fill'], Font(bold=True, color="000000"), styles['thin_border']
        
        for item in cat_items:
            unit_price_inr = item.get('price', 0) * usd_to_inr_rate
            subtotal = unit_price_inr * item.get('quantity', 1)
            gst_rate = item.get('gst_rate', gst_rates.get('Electronics', 18))
            sgst_rate = cgst_rate = gst_rate / 2
            sgst_amount, cgst_amount = subtotal * (sgst_rate / 100), subtotal * (cgst_rate / 100)
            total_tax = sgst_amount + cgst_amount
            total_with_gst = subtotal + total_tax
            total_before_gst_hardware += subtotal
            row_data = [item_s_no, item.get('name', ''), item.get('brand', 'Unknown'), item.get('model_number', 'N/A'), item.get('quantity', 1), unit_price_inr, subtotal, item.get('warranty', 'Not Specified'), f"{sgst_rate}%", sgst_amount, f"{cgst_rate}%", cgst_amount, total_tax, total_with_gst, '']
            sheet.append(row_data)
            current_row = sheet.max_row
            
            if IMAGE_GENERATION_AVAILABLE:
                try:
                    size_inches = extract_display_size(item.get('name', '')) if item.get('category') == 'Displays' else None
                    img_buffer = generate_product_info_card(product_name=item.get('name', 'Unknown Product'), brand=item.get('brand', 'N/A'), model=item.get('model_number', 'N/A'), category=item.get('category', 'General AV'), size_inches=size_inches)
                    if img_buffer:
                        img_buffer.seek(0)
                        excel_img = ExcelImage(img_buffer)
                        excel_img.width, excel_img.height = 150, 100
                        sheet.add_image(excel_img, f'O{current_row}')
                        sheet.row_dimensions[current_row].height = 85
                except Exception as e: print(f"⚠️ Could not add product image for {item.get('name', 'Unknown')}: {e}")
            item_s_no += 1

    # Add Services
    services_gst_rate = gst_rates.get('Services', 18)
    if total_before_gst_hardware > 0:
        sheet.append([chr(ord('A') + len(grouped_items)), "Services"])
        cat_row_idx = sheet.max_row
        sheet.merge_cells(f'B{cat_row_idx}:O{cat_row_idx}')
        for cell in sheet[cat_row_idx]: cell.fill, cell.font, cell.border = styles['boq_category_fill'], styles['bold_font'], styles['thin_border']
        for service_name, percentage in [("Installation & Commissioning", 0.15), ("System Warranty (3 Years)", 0.05), ("Project Management", 0.10)]:
            service_amount_inr = total_before_gst_hardware * percentage
            sgst_rate = cgst_rate = services_gst_rate / 2
            service_sgst, service_cgst = service_amount_inr * (sgst_rate / 100), service_amount_inr * (cgst_rate / 100)
            service_total = service_amount_inr + service_sgst + service_cgst
            row_data = [item_s_no, service_name, "AllWave AV", "Professional Service", 1, service_amount_inr, service_amount_inr, "As per terms", f"{sgst_rate}%", service_sgst, f"{cgst_rate}%", service_cgst, service_sgst + service_cgst, service_total, '']
            sheet.append(row_data)
            item_s_no += 1
    
    # Column Widths and Final Formatting
    widths = {'A': 8, 'B': 45, 'C': 20, 'D': 30, 'E': 6, 'F': 15, 'G': 15, 'H': 15, 'I': 10, 'J': 15, 'K': 10, 'L': 15, 'M': 15, 'N': 18, 'O': 25}
    for col, width in widths.items(): sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows(min_row=header_start_row + 2, max_col=15):
        for cell in row:
            if isinstance(cell, MergedCell): continue
            if cell.column in [6, 7, 10, 12, 13, 14] and isinstance(cell.value, (int, float)): cell.number_format = styles['currency_format']
            cell.border = styles['thin_border']
            cell.alignment = Alignment(horizontal='center', vertical='top') if cell.column in [1, 5] else Alignment(vertical='top', wrap_text=True)


# ==================== SCOPE OF WORK SHEET (FINAL CONTENT) ====================
def _add_scope_of_work_sheet(workbook, styles):
    """Creates the Scope of Work sheet with all static content."""
    sheet = workbook.create_sheet(title="Scope of Work", index=2)
    sheet.sheet_view.showGridLines = False
    _create_sheet_header(sheet)
    sheet.column_dimensions['A'].width, sheet.column_dimensions['B'].width = 8, 80
    row_cursor = 4

    # Intro Paragraph
    sheet.merge_cells(f'A{row_cursor}:B{row_cursor}')
    intro_cell = sheet[f'A{row_cursor}']
    intro_cell.value = "This document outlines the scope of work for the proposed Audio Visual Solution. It details the responsibilities of All Wave AV Systems and the dependencies on the client to ensure a smooth and successful project implementation."
    intro_cell.alignment = Alignment(wrap_text=True, vertical='top')
    sheet.row_dimensions[row_cursor].height = 60
    row_cursor += 2

    # Project Phases
    sheet.merge_cells(f'A{row_cursor}:B{row_cursor}')
    phases_header = sheet[f'A{row_cursor}']
    phases_header.value, phases_header.font, phases_header.fill = "PROJECT PHASES", styles['bold_font'], styles['header_light_green_fill']
    phases_header.border, sheet[f'B{row_cursor}'].border = styles['thin_border'], styles['thin_border']
    row_cursor += 1
    phases = [
        ("Design", "Detailed engineering drawings, schematics, and equipment layouts."),
        ("Procurement", "Ordering and consolidation of all specified equipment."),
        ("Implementation", "Installation, cabling, configuration, and programming of the AV system."),
        ("Handover", "System commissioning, user training, and final documentation.")
    ]
    for label, desc in phases:
        sheet[f'A{row_cursor}'].value, sheet[f'A{row_cursor}'].font, sheet[f'A{row_cursor}'].border = label, styles['bold_font'], styles['thin_border']
        sheet[f'B{row_cursor}'].value, sheet[f'B{row_cursor}'].border = desc, styles['thin_border']
        row_cursor += 1
    row_cursor += 2

    # Main Title
    sheet.merge_cells(f'A{row_cursor}:B{row_cursor}')
    title_cell = sheet[f'A{row_cursor}']
    title_cell.value, title_cell.font, title_cell.fill = "Scope of Work", Font(size=14, bold=True, color="FFFFFF"), PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    title_cell.alignment, title_cell.border = Alignment(horizontal='center', vertical='center'), styles['thin_border']
    sheet.row_dimensions[row_cursor].height = 25
    row_cursor += 2
    
    scope_items = ["Site Coordination and Prerequisites Clearance", "Detailed schematic drawings according to the design", "Conduit layout drawings/equipment layout drawings, showing mounting location", "Laying of all AV Cables", "Termination of cables with respective connectors", "Installation of all AV equipment in rack as per layout", "Configuration of Audio/Video Switcher", "Configuration of DSP mixer", "Touch Panel Design", "System programming as per design requirement", "As built drawings after completion of project", "User training & Handover"]
    for idx, item in enumerate(scope_items, 1):
        sheet[f'A{row_cursor}'].value, sheet[f'A{row_cursor}'].alignment, sheet[f'A{row_cursor}'].border, sheet[f'A{row_cursor}'].fill = idx, Alignment(horizontal='center', vertical='center'), styles['thin_border'], styles['header_light_green_fill']
        sheet[f'B{row_cursor}'].value, sheet[f'B{row_cursor}'].border, sheet[f'B{row_cursor}'].alignment = item, styles['thin_border'], Alignment(vertical='center', wrap_text=True)
        sheet.row_dimensions[row_cursor].height = 30
        row_cursor += 1
    row_cursor += 1
    
    # Exclusions Section
    sheet.merge_cells(f'A{row_cursor}:B{row_cursor}')
    section_cell = sheet[f'A{row_cursor}']
    section_cell.value, section_cell.fill, section_cell.font = "Exclusions and Dependencies", styles['table_header_blue_fill'], Font(bold=True, color="FFFFFF")
    section_cell.alignment, section_cell.border, sheet[f'B{row_cursor}'].border = Alignment(horizontal='left', vertical='center'), styles['thin_border'], styles['thin_border']
    row_cursor += 1
    sheet.merge_cells(f'A{row_cursor}:B{row_cursor}')
    cell = sheet[f'A{row_cursor}']
    cell.value, cell.border, cell.alignment = "The following items need to be arranged by the client on site:", styles['thin_border'], Alignment(vertical='center')
    sheet[f'B{row_cursor}'].border, sheet.row_dimensions[row_cursor].height = styles['thin_border'], 20
    row_cursor += 1
    
    exclusions = ["Civil work like cutting of false ceilings, chipping, etc.", "Electrical work like laying of conduits, raceways, and providing stabilized power supply", "Carpentry work like cutouts on furniture, etc.", "Connectivity for electric power, LAN, telephone, IP (1 Mbps), ISDN (1 Mbps) & cable TV points", "Ballasts (0 to 10 volts) in case of fluorescent dimming for lights", "Shelves for mounting devices (if a rack is not in the SOW)", "Adequate cooling/ventilation for all equipment racks and cabinets", "Any software other than specified in the proposal", "Any active LAN points required for equipment"]
    for idx, item in enumerate(exclusions, 1):
        sheet[f'A{row_cursor}'].value, sheet[f'A{row_cursor}'].alignment, sheet[f'A{row_cursor}'].border, sheet[f'A{row_cursor}'].fill = idx, Alignment(horizontal='center', vertical='center'), styles['thin_border'], styles['header_light_green_fill']
        sheet[f'B{row_cursor}'].value, sheet[f'B{row_cursor}'].border, sheet[f'B{row_cursor}'].alignment = item, styles['thin_border'], Alignment(wrap_text=True, vertical='center')
        sheet.row_dimensions[row_cursor].height = 25
        row_cursor += 1


# ==================== PROPOSAL SUMMARY SHEET (FINAL) ====================
def _add_proposal_summary_sheet(workbook, rooms_data, project_details, styles):
    """ Creates the Proposal Summary sheet with highlights. """
    sheet = workbook.create_sheet(title="Proposal Summary", index=3)
    sheet.sheet_view.showGridLines = False
    _create_sheet_header(sheet)
    widths = {'A': 10, 'B': 50, 'C': 12, 'D': 18, 'E': 18, 'F': 18, 'G': 18}
    for col, width in widths.items(): sheet.column_dimensions[col].width = width
    row_cursor = 4

    # Title and Headers
    sheet.merge_cells(f'A{row_cursor}:G{row_cursor}')
    title_cell = sheet[f'A{row_cursor}']
    title_cell.value, title_cell.font, title_cell.fill = "Proposal Summary", Font(size=14, bold=True, color="FFFFFF"), PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    title_cell.alignment, title_cell.border = Alignment(horizontal='center', vertical='center'), styles['thin_border']
    sheet.row_dimensions[row_cursor].height = 25
    row_cursor += 2
    headers1, headers2 = ['Sr. No', 'Description', 'Total Qty', '', 'INR Supply', '', ''], ['', '', '', 'Rate w/o TAX', 'Amount w/o TAX', 'Total TAX Amount', 'Amount with Tax']
    for col_idx, header in enumerate(headers1, 1):
        cell = sheet.cell(row=row_cursor, column=col_idx)
        cell.value, cell.fill, cell.font, cell.alignment, cell.border = header, styles['table_header_blue_fill'], Font(bold=True, color="FFFFFF"), Alignment(horizontal='center', vertical='center'), styles['thin_border']
    sheet.merge_cells(f'D{row_cursor}:G{row_cursor}')
    row_cursor += 1
    for col_idx, header in enumerate(headers2, 1):
        cell = sheet.cell(row=row_cursor, column=col_idx)
        cell.value, cell.fill, cell.font, cell.alignment, cell.border = header, styles['table_header_blue_fill'], Font(bold=True, color="FFFFFF"), Alignment(horizontal='center', vertical='center', wrap_text=True), styles['thin_border']
    sheet.row_dimensions[row_cursor].height = 30
    row_cursor += 1

    # Room Data
    grand_subtotal, grand_tax, grand_total = 0, 0, 0
    for idx, room in enumerate(rooms_data, 1):
        room_subtotal, room_tax, room_total = room.get('subtotal', 0), room.get('gst', 0), room.get('total', 0)
        grand_subtotal += room_subtotal
        grand_tax += room_tax
        grand_total += room_total
        total_qty = sum(item.get('quantity', 1) for item in room.get('boq_items', []))
        avg_rate = room_subtotal / total_qty if total_qty > 0 else 0
        row_data = [idx, room.get('name', f'Room {idx}'), total_qty, avg_rate, room_subtotal, room_tax, room_total]
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_cursor, column=col_idx)
            cell.value, cell.border = value, styles['thin_border']
            if col_idx >= 4: cell.number_format, cell.alignment = styles['currency_format'], Alignment(horizontal='right', vertical='center')
            elif col_idx in [1, 3]: cell.alignment = Alignment(horizontal='center', vertical='center')
            else: cell.alignment = Alignment(horizontal='left', vertical='center')
        sheet.row_dimensions[row_cursor].height = 20
        row_cursor += 1
    
    # Grand Total Row
    row_cursor += 1
    sheet.merge_cells(f'A{row_cursor}:C{row_cursor}')
    total_label_cell = sheet[f'A{row_cursor}']
    total_label_cell.value, total_label_cell.font, total_label_cell.fill = "GRAND TOTAL", Font(bold=True, size=12), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    total_label_cell.alignment, total_label_cell.border = Alignment(horizontal='center', vertical='center'), styles['thin_border']
    grand_total_data = ['', grand_subtotal, grand_tax, grand_total]
    for col_idx, value in enumerate(grand_total_data, 4):
        cell = sheet.cell(row=row_cursor, column=col_idx)
        cell.value, cell.font, cell.fill, cell.border = value, Font(bold=True, size=11), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), styles['thin_border']
        if value: cell.number_format, cell.alignment = styles['currency_format'], Alignment(horizontal='right', vertical='center')
    sheet.row_dimensions[row_cursor].height = 25
    row_cursor += 3

    # Project Metadata Highlights
    if project_details.get('PSNI Referral') == 'Yes':
        sheet.merge_cells(f'A{row_cursor}:G{row_cursor}')
        psni_cell = sheet[f'A{row_cursor}']
        psni_cell.value = "✅ PSNI GLOBAL ALLIANCE REFERRED PROJECT"
        psni_cell.font, psni_cell.fill = Font(size=12, bold=True, color="FFFFFF"), PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        psni_cell.alignment, psni_cell.border, sheet.row_dimensions[row_cursor].height = Alignment(horizontal='center', vertical='center'), styles['thin_border'], 25
        row_cursor += 1

    client_type = project_details.get('Client Type', 'International')
    color = "3B82F6" if client_type == "Local (India)" else "8B5CF6"
    sheet.merge_cells(f'A{row_cursor}:G{row_cursor}')
    client_cell = sheet[f'A{row_cursor}']
    client_cell.value, client_cell.font, client_cell.fill = f"🌐 CLIENT TYPE: {client_type.upper()}", Font(size=11, bold=True, color="FFFFFF"), PatternFill(start_color=color, end_color=color, fill_type="solid")
    client_cell.alignment, client_cell.border, sheet.row_dimensions[row_cursor].height = Alignment(horizontal='center', vertical='center'), styles['thin_border'], 22


# ==================== EXECUTIVE SUMMARY SHEET ====================
def generate_budget_summary_sheet(workbook, rooms_data, project_details, styles):
    """ Create executive-level budget summary (1-page overview) """
    sheet = workbook.create_sheet(title="Executive Summary", index=1)
    sheet.sheet_view.showGridLines = False
    _create_sheet_header(sheet)
    row = 4
    
    # Project Overview
    sheet.merge_cells(f'A{row}:F{row}')
    header = sheet[f'A{row}']
    header.value, header.fill, header.font, header.alignment = "PROJECT OVERVIEW", styles['table_header_blue_fill'], Font(bold=True, color="FFFFFF"), Alignment(horizontal='center', vertical='center')
    row += 2
    overview_data = [("Project Name", project_details.get('Project Name', 'N/A')), ("Client", project_details.get('Client Name', 'N/A')), ("Location", project_details.get('Location', 'N/A')), ("Total Rooms", len(rooms_data)), ("Project Date", datetime.now().strftime("%B %d, %Y"))]
    for label, value in overview_data:
        sheet[f'A{row}'].value, sheet[f'A{row}'].font = label, styles['bold_font']
        sheet.merge_cells(f'B{row}:F{row}')
        sheet[f'B{row}'].value = value
        row += 1
    row += 2
    
    # Budget Summary by Room
    sheet.merge_cells(f'A{row}:F{row}')
    header = sheet[f'A{row}']
    header.value, header.fill, header.font, header.alignment = "BUDGET BREAKDOWN BY SPACE", styles['table_header_blue_fill'], Font(bold=True, color="FFFFFF"), Alignment(horizontal='center', vertical='center')
    row += 1
    headers = ['Room Name', 'Area (sqft)', 'Equipment Cost', 'Services', 'Tax', 'Total']
    for col_idx, header_text in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=col_idx)
        cell.value, cell.fill, cell.font, cell.border, cell.alignment = header_text, styles['header_light_green_fill'], styles['bold_font'], styles['thin_border'], Alignment(horizontal='center', vertical='center')
    row += 1
    
    grand_equipment, grand_services, grand_tax, grand_total = 0, 0, 0, 0
    for room in rooms_data:
        equipment_cost, services_cost = room.get('subtotal', 0) / 1.30, (room.get('subtotal', 0) / 1.30) * 0.30
        tax, total = room.get('gst', 0), room.get('total', 0)
        grand_equipment += equipment_cost; grand_services += services_cost; grand_tax += tax; grand_total += total
        room_data = [room.get('name', 'Unknown'), f"{room.get('area', 0):.0f}", equipment_cost, services_cost, tax, total]
        for col_idx, value in enumerate(room_data, 1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value, cell.border = value, styles['thin_border']
            cell.alignment = Alignment(horizontal='right', vertical='center') if col_idx >= 3 else Alignment(horizontal='center', vertical='center')
            if col_idx >= 3: cell.number_format = styles['currency_format']
        row += 1
    
    # Grand total row
    sheet.merge_cells(f'A{row}:B{row}')
    total_label = sheet[f'A{row}']
    total_label.value, total_label.font, total_label.fill, total_label.border = "TOTAL PROJECT INVESTMENT", Font(bold=True, size=12), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), styles['thin_border']
    sheet[f'B{row}'].border = styles['thin_border']
    totals = [grand_equipment, grand_services, grand_tax, grand_total]
    for col_idx, value in enumerate(totals, 3):
        cell = sheet.cell(row=row, column=col_idx)
        if value is not None: cell.value, cell.number_format, cell.alignment = value, styles['currency_format'], Alignment(horizontal='right', vertical='center')
        cell.font, cell.fill, cell.border = Font(bold=True, size=11), PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), styles['thin_border']
    widths = {'A': 30, 'B': 15, 'C': 18, 'D': 18, 'E': 18, 'F': 20}
    for col, width in widths.items(): sheet.column_dimensions[col].width = width


# ==================== MAIN ENTRY POINT ====================
def generate_company_excel(project_details, rooms_data, usd_to_inr_rate):
    """ Main function to generate the complete Excel workbook. """
    workbook = openpyxl.Workbook()
    styles = _define_styles()

    # Calculate room totals first
    for room in rooms_data:
        if room.get('boq_items'):
            subtotal_hardware = sum(item.get('price', 0) * item.get('quantity', 1) for item in room['boq_items']) * usd_to_inr_rate
            services_total = subtotal_hardware * 0.30
            total_without_gst = subtotal_hardware + services_total
            gst_electronics = sum((item.get('price', 0) * item.get('quantity', 1) * usd_to_inr_rate) * (item.get('gst_rate', 18) / 100) for item in room['boq_items'])
            gst_services = services_total * (project_details.get('gst_rates', {}).get('Services', 18) / 100)
            total_gst = gst_electronics + gst_services
            room['subtotal'], room['gst'], room['total'] = total_without_gst, total_gst, total_without_gst + total_gst

    # Generate sheets in order
    _add_version_control_sheet(workbook, project_details, styles)
    generate_budget_summary_sheet(workbook, rooms_data, project_details, styles)
    _add_scope_of_work_sheet(workbook, styles)
    _add_proposal_summary_sheet(workbook, rooms_data, project_details, styles)
    _add_terms_and_conditions_sheet(workbook, styles)
    
    # Add a BOQ sheet for each room with items
    for room in rooms_data:
        if room.get('boq_items'):
            safe_name = re.sub(r'[\\/*?:"<>|]', '', room['name'])[:25]
            room_sheet = workbook.create_sheet(title=f"BOQ - {safe_name}")
            _populate_room_boq_sheet(room_sheet, room['boq_items'], room['name'], styles, usd_to_inr_rate, project_details.get('gst_rates', {}))

    # Cleanup and save
    if "Sheet" in workbook.sheetnames: del workbook["Sheet"]
    workbook.active = workbook["Version Control"]
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()
